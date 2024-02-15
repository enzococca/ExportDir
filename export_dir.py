import tempfile
import sys
import time
from PyQt5.QtGui import QPixmap

from PyQt5.QtMultimedia import QSound

from PyQt5.QtCore import Qt, QTimer, QThread, pyqtSignal
from PyQt5.QtWidgets import QFileDialog, QApplication, QProgressDialog,QMessageBox,QSplashScreen
import os
from openpyxl import Workbook
from openpyxl import styles
from urllib.request import pathname2url
from comtypes.client import CreateObject
import getpass
import locale



class DirectoryToExcelAndPdfConverter:
    def __init__(self):
        self.app = QApplication(sys.argv)

        # Splash screen setup

        splash_pix = QPixmap('pdf.png')  # Replace with the path to your splash image
        splash = QSplashScreen(splash_pix)
        splash.show()
        QApplication.processEvents()  # Process pending events to display the splash screen immediately
        QSound.play('intro.wav')  # Replace with the path to your sound file
        # Setup a timer to close the splash screen after 3 seconds
        # Setup a timer to close the splash screen after 3 seconds and show the welcome message
        QTimer.singleShot(3000, lambda: (splash.close(), self.show_welcome_message()))

        self.root_path = ''
        self.output_excel_file_1 = ''
        self.output_excel_file_2 = ''
        self.username = getpass.getuser()
        self.default_locale = locale.getdefaultlocale()[0]
        self.setup_messages()


    def setup_messages(self):
        # Define other messages based on the system's locale
        self.messages = {
            'it': {
                'select_root': "Seleziona la cartella radice",
                'select_first_excel': "Seleziona il primo file PDF di output riguardante directory e file",
                'select_second_excel': "Seleziona il secondo file PDF di output riguardante solo la directory",
                'operation_in_progress': "Operazione in corso...",
                'operation_completed': "Operazione completata con successo."
            },
            'es': {
                'select_root': "Seleccionar directorio raíz",
                'select_first_excel': "Seleccionar el primer archivo PDF de salida sobre directorio y archivos",
                'select_second_excel': "Seleccionar el segundo archivo PDF de salida solo sobre directorio",
                'operation_in_progress': "Operación en curso...",
                'operation_completed': "Operación completada con éxito."
            },
            'fr': {
                'select_root': "Sélectionner le répertoire racine",
                'select_first_excel': "Sélectionnez le premier fichier PDF de sortie concernant le répertoire et les fichiers",
                'select_second_excel': "Sélectionnez le deuxième fichier PDF de sortie concernant uniquement le répertoire",
                'operation_in_progress': "Opération en cours...",
                'operation_completed': "Opération terminée avec succès."
            },

            'default': {
                'select_root': "Select Root Directory",
                'select_first_excel': "Select First Output PDF File about directory and files",
                'select_second_excel': "Select Second Output PDF File about directory only",
                'operation_in_progress': "Operation in progress...",
                'operation_completed': "Operation completed successfully."
            }
        }
        # Get the current username from the operating system
        self.username = getpass.getuser()


        self.current_messages = self.messages.get(self.default_locale[:2], self.messages['default'])

    def show_welcome_message(self):
        # Determine the greeting message based on the system's locale
        greetings = {
            'it': f"Benvenuto {self.username} nel programma di esportazione della struttura delle directory\n"
                  "Scegli una cartella radice e poi salva due file PDF in output.",
            'es': f"Bienvenido {self.username} al programa de exportación de estructura de directorios.\n"
                  "Elige un directorio raíz y luego guarda dos archivos PDF de salida.",
            'fr': f"Bienvenue {self.username} dans le programme d'exportation de la structure des répertoires.\n"
                  "Choisissez un répertoire racine et enregistrez deux fichiers PDF en sortie.",
            'default': f"Welcome {self.username} to the Export directory structure program.\n"
                       "Choose a root folder and then save two output PDF files."
        }
        greeting = greetings.get(self.default_locale[:2], greetings['default'])
        # Show the message box and connect the button click to the next action
        self.msg_box = QMessageBox()
        self.msg_box.setWindowTitle('Export Directory Structure')
        self.msg_box.setText(greeting)
        self.msg_box.addButton(QMessageBox.Ok)
        self.msg_box.buttonClicked.connect(self.on_welcome_message_closed)
        self.msg_box.exec_()  # Blocking call until the user closes the message box

    def on_welcome_message_closed(self, event):
        print(self.output_excel_file_1, self.output_excel_file_2)
        if self.select_root_path() and self.select_output_excel_paths():
            self.create_pdfs_from_directory()


    def select_root_path(self):
        title = self.current_messages['select_root']
        self.root_path = QFileDialog.getExistingDirectory(None, title)
        if self.root_path:  # Proceed only if a directory was selected
            self.select_output_pdf_paths()

    def select_output_pdf_paths(self):
        first_title = self.current_messages['select_first_excel']
        second_title = self.current_messages['select_second_excel']

        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog

        # Select the first PDF file
        self.output_pdf_file_1, _ = QFileDialog.getSaveFileName(None, first_title, "", "PDF Files (*.pdf)",
                                                                options=options)
        if not self.output_pdf_file_1:
            return

        # Select the second PDF file
        self.output_pdf_file_2, _ = QFileDialog.getSaveFileName(None, second_title, "", "PDF Files (*.pdf)",
                                                                options=options)
        if self.output_pdf_file_1 and self.output_pdf_file_2:
            self.create_pdfs_from_directory()

    def add_directory_to_excel(self, ws, path, indent=0, font_size=18):
        items = sorted(os.listdir(path))
        for item in items:
            item_path = os.path.join(path, item)
            display_name = f"{'   ' * indent}{item}"
            ws.append([display_name])
            cell = ws.cell(row=ws.max_row, column=1)
            cell.hyperlink = self.create_file_hyperlink(item_path)
            cell.font = styles.Font(bold=True, size=font_size)
            if os.path.isdir(item_path):
                new_font_size = max(8, font_size - 2)
                self.add_directory_to_excel(ws, item_path, indent + 2, new_font_size)

    def add_directories_to_excel(self, ws, path, indent=0, font_size=18):
        with os.scandir(path) as it:
            for entry in sorted(it, key=lambda e: e.name):
                if entry.is_dir():
                    display_name = f"{'   ' * indent}.{entry.name}"
                    ws.append([display_name])
                    cell = ws.cell(row=ws.max_row, column=1)
                    cell.hyperlink = self.create_file_hyperlink(entry.path)
                    cell.font = styles.Font(bold=True, size=font_size)
                    new_font_size = max(8, font_size - 2)
                    self.add_directories_to_excel(ws, entry.path, indent + 4, new_font_size)

    def create_file_hyperlink(self, path):
        # Convert the path to an absolute path
        abs_path = os.path.abspath(path)
        # Create a file URI that can be used as a hyperlink in PDF
        file_uri = 'file:///' + pathname2url(abs_path)
        return file_uri

    def select_output_excel_paths(self):
        first_title = self.current_messages['select_first_excel']
        second_title = self.current_messages['select_second_excel']

        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog

        # Prompt the user to select the first Excel file path
        self.output_excel_file_1, _ = QFileDialog.getSaveFileName(
            None, first_title, "", "Excel Files (*.xlsx)", options=options)
        if not self.output_excel_file_1:
            return False  # User cancelled or closed the dialog

        # Ensure the file has the correct .xlsx extension
        if not self.output_excel_file_1.endswith('.xlsx'):
            self.output_excel_file_1 += '.xlsx'

        # Prompt the user to select the second Excel file path
        self.output_excel_file_2, _ = QFileDialog.getSaveFileName(
            None, second_title, "", "Excel Files (*.xlsx)", options=options)
        if not self.output_excel_file_2:
            return False  # User cancelled or closed the dialog

        # Ensure the file has the correct .xlsx extension
        if not self.output_excel_file_2.endswith('.xlsx'):
            self.output_excel_file_2 += '.xlsx'
        print(self.output_excel_file_1,self.output_excel_file_2)
        return True  # Both files were selected

    def start_excel_creation_process(self):
        print(self.output_excel_file_1, self.output_excel_file_2)
        # Start the process only if valid file paths are provided
        if self.root_path and self.output_excel_file_1 and self.output_excel_file_2:
            self.create_excels_from_directory()
        else:
            QMessageBox.warning(None, "Error", "Invalid file paths provided.")

    



    def create_pdfs_from_directory(self):
        if self.root_path and self.output_pdf_file_1 and self.output_pdf_file_2:
            in_progress = self.current_messages['operation_in_progress']
            completed = self.current_messages['operation_completed']

            self.progress_dialog = QProgressDialog(in_progress, None, 0, 2)
            self.progress_dialog.setWindowModality(Qt.WindowModal)
            self.progress_dialog.setAutoClose(False)
            self.progress_dialog.setMinimumDuration(0)
            self.progress_dialog.show()  # Show the progress dialog immediately
            QApplication.processEvents()  # Process events to ensure the dialog is displayed


            self.worker_thread = WorkerThread(self)
            self.worker_thread.update_progress.connect(self.update_progress_dialog)
            self.worker_thread.completed.connect(self.on_worker_thread_finished)
            self.worker_thread.start()

    def update_progress_dialog(self, value):
        self.progress_dialog.setValue(value)

    def create_excels_from_directory(self):
        # Show progress dialog and start worker thread
        self.progress_dialog = QProgressDialog("Operation in progress...", "Cancel", 0, 100)
        self.progress_dialog.setWindowModality(Qt.WindowModal)
        self.progress_dialog.setAutoClose(False)
        self.progress_dialog.setMinimumDuration(0)
        self.progress_dialog.show()
        QApplication.processEvents()

        self.worker_thread = WorkerThread(self)
        self.worker_thread.update_progress.connect(self.progress_dialog.setValue)
        self.worker_thread.completed.connect(self.on_worker_thread_finished)  # Connect the slot to the signal
        self.worker_thread.start()

    def on_worker_thread_finished(self, elapsed_time, temp_excel_file_1, temp_excel_file_2):
        # Convert the temporary Excel files to PDF in the main thread
        self.convert_excel_to_pdf(temp_excel_file_1, self.output_pdf_file_1)
        self.convert_excel_to_pdf(temp_excel_file_2, self.output_pdf_file_2)

        # Update the progress dialog with the elapsed time
        self.progress_dialog.setLabelText(f"Operation completed in {elapsed_time:.2f} seconds.")
        self.progress_dialog.setCancelButtonText("Close")
        self.progress_dialog.canceled.connect(self.progress_dialog.close)
        self.progress_dialog.setValue(self.progress_dialog.maximum())

        # Delete the temporary Excel files
        os.unlink(temp_excel_file_1)
        os.unlink(temp_excel_file_2)

    def convert_excel_to_pdf(self, excel_file, output_pdf_file):
        try:
            excel = CreateObject("Excel.Application")  # Corrected ProgID for Excel
            wb = excel.Workbooks.Open(excel_file)
            ws = wb.ActiveSheet
            ws.ExportAsFixedFormat(0, output_pdf_file)  # 0 stands for PDF format
            wb.Close(False)
            excel.Quit()
        except Exception as e:
            print(f"Error converting {excel_file} to PDF: {e}")
class WorkerThread(QThread):
    update_progress = pyqtSignal(int)
    completed = pyqtSignal(float,str,str)  # Signal to indicate completion and pass the temporary file paths

    def __init__(self, converter):
        super().__init__()
        self.converter = converter

    def run(self):
        start_time=time.time()
        # Create temporary Excel files
        temp_excel_file_1 = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_excel_file_2 = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')

        try:
            # Create the first Excel file and save it to a temporary path
            wb1 = Workbook()
            ws1 = wb1.active
            self.converter.add_directory_to_excel(ws1, self.converter.root_path)
            wb1.save(temp_excel_file_1.name)
            self.update_progress.emit(50)

            # Create the second Excel file and save it to a temporary path
            wb2 = Workbook()
            ws2 = wb2.active
            self.converter.add_directories_to_excel(ws2, self.converter.root_path)
            wb2.save(temp_excel_file_2.name)
            self.update_progress.emit(100)

            elapsed_time = time.time() - start_time
            self.completed.emit(elapsed_time, temp_excel_file_1.name, temp_excel_file_2.name)
        finally:
            # Close the temporary files
            temp_excel_file_1.close()
            temp_excel_file_2.close()


# Example usage:
if __name__ == "__main__":
    converter = DirectoryToExcelAndPdfConverter()
    sys.exit(converter.app.exec_())


